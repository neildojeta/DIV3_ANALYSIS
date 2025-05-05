import xlwings as xw
import openpyxl
from xlwings.utils import rgb_to_int
import win32com.client
from openpyxl.drawing.image import Image
from PIL import Image as PILImage, ImageDraw, ImageFont
import os
import logging
import time
import sys

# Set up logging
log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(log_folder, 'Comparison.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Define the comparison files and corresponding sheets
def main(file_previous, file_latest, prev_operator, prev_date, lat_operator, lat_date): 
    logger.info(f"{file_previous} + {file_latest}")
    comparison_files = [
        ('ComparedResults/DIV3_Main_Tables.xlsx', 'Dashboard'),
        ('ComparedResults/DIV3_ADA_Tables.xlsx', 'ADA'),
        ('ComparedResults/DIV3_GOLINK_Tables.xlsx', 'GOLINK'),
        ('ComparedResults/DIV3_ADAGOLINK_Tables.xlsx', 'ADAGOLINK')
    ]

    # Open the Dashboard workbook
    dashboard_file = 'ComparedResults/Dashboard.xlsm'
    app = xw.App(visible=True)

    try:
        # Open the existing workbook (Dashboard)
        wb_dashboard = app.books.open(dashboard_file)

        # Loop through each comparison file and corresponding sheet
        for comparison_file, sheet_name in comparison_files:
            # Load the comparison workbook and sheet
            wb_comparison = openpyxl.load_workbook(comparison_file)
            
            # sheet_OperatorComparison = wb_comparison['OperatorChanges']
            # sheet_DatesComparison = wb_comparison['MissingDates']


            if sheet_name == 'Dashboard':
                sheet_TotalInvoicePayment = wb_comparison['TotalInvoicePayment']
                sheet_TTLRevComparison = wb_comparison['HourlyTTLRevComparison']
                sheet_LiftLeaseComparison = wb_comparison['LiftLeaseComparison']
                sheet_ViolationComparison = wb_comparison['ViolationsComparison']
                sheet_CashCollectedComparison = wb_comparison['CashCollectedComparison']
                # Retrieve Payment values from the 'TotalInvoicePayment' sheet in the comparison file
                full_prev_amount = f"{float(sheet_TotalInvoicePayment['B2'].value):,.2f}"
                full_lat_amount = f"{float(sheet_TotalInvoicePayment['A2'].value):,.2f}"

                # Retrieve values from the 'TotalInvoicePayment' sheet in the comparison file
                full_amount_diff = f"{abs(float(sheet_TotalInvoicePayment['C2'].value)):,.2f}"
                full_amount_diff_status = sheet_TotalInvoicePayment['D2'].value

                # Retrieve values from the 'HourlyTTLRevComparison' sheet in the comparison file
                lat_HTTLRevComp = f"{sum(cell.value for row in sheet_TTLRevComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_HTTLRevComp = f"{sum(cell.value for row in sheet_TTLRevComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_HTTLRevComp = f"{sum(cell.value for row in sheet_TTLRevComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                # Retrieve values from the 'LiftLeaseComparison' sheet in the comparison file
                lat_LeaseComp = f"{sum(cell.value for row in sheet_LiftLeaseComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_LeaseComp = f"{sum(cell.value for row in sheet_LiftLeaseComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_LeaseComp = f"{sum(cell.value for row in sheet_LiftLeaseComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                # Retrieve values from the 'ViolationComparison' sheet in the comparison file
                lat_ViolationComp = f"{sum(cell.value for row in sheet_ViolationComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_ViolationComp = f"{sum(cell.value for row in sheet_ViolationComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_ViolationComp = f"{sum(cell.value for row in sheet_ViolationComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                lat_CashCollectedComp = f"{sum(cell.value for row in sheet_CashCollectedComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_CashCollectedComp = f"{sum(cell.value for row in sheet_CashCollectedComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_CashCollectedComp = f"{sum(cell.value for row in sheet_CashCollectedComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                # # Retrieve values from the 'OperatorChanges' sheet in the comparison file
                # prev_OperatorChanges = prev_operator
                # lat_OperatorChanges = lat_operator
                # diff_OperatorChanges = lat_operator - prev_operator
                
                # # Retrieve values from the 'MissingDates' sheet in the comparison file
                # prev_DateChanges = prev_date
                # lat_DateChanges = lat_date
                # diff_DateChanges = lat_date - prev_date

                # Get the corresponding sheet in the dashboard
                sheet_dashboard = wb_dashboard.sheets[sheet_name]

                # Access the file name values shape via the API and set the value
                txt_prev_file = sheet_dashboard.shapes['txtPrevFile'].api
                txt_prev_file.TextFrame2.TextRange.Text = f"{file_previous}"

                txt_lat_file = sheet_dashboard.shapes['txtLatFile'].api
                txt_lat_file.TextFrame2.TextRange.Text = f"{file_latest}"

                # Access the total diff value shape via the API and set the value
                txt_full_amount_diff = sheet_dashboard.shapes['TextBox 88'].api
                txt_full_amount_diff.TextFrame2.TextRange.Text = f"$ {full_amount_diff}"

                txt_full_prev_amount = sheet_dashboard.shapes['txtPrevPAyment'].api
                txt_full_prev_amount.TextFrame2.TextRange.Text = f"$ {full_prev_amount}"

                txt_full_lat_amount = sheet_dashboard.shapes['txtLatPayment'].api
                txt_full_lat_amount.TextFrame2.TextRange.Text = f"$ {full_lat_amount}"

                # Access the total diff status shape via the API and set the value
                txt_full_amount_diff_status = sheet_dashboard.shapes['TextBox 90'].api
                txt_full_amount_diff_status.TextFrame2.TextRange.Text = f"{full_amount_diff_status}"

                #For HTTL Rev
                txt_HTTLRevComp = sheet_dashboard.shapes['txtDHTTLRevtDiff'].api
                txt_HTTLRevComp.TextFrame2.TextRange.Text = f"${prev_HTTLRevComp} to ${lat_HTTLRevComp}"
                txt_HTTLRevComp_diff = sheet_dashboard.shapes['txtHTTLRevtDiff'].api
                txt_HTTLRevComp_diff.TextFrame2.TextRange.Text = f"${diff_HTTLRevComp}"

                # Access the LiftLeaseComparison shape via the API and set the value
                txt_LiftLease = sheet_dashboard.shapes['txtDLLeaseDiff'].api
                txt_LiftLease.TextFrame2.TextRange.Text = f"${prev_LeaseComp} to ${lat_LeaseComp}"
                txt_LiftLease_diff = sheet_dashboard.shapes['txtLLeaseDiff'].api
                txt_LiftLease_diff.TextFrame2.TextRange.Text = f"${diff_LeaseComp}"

                # Access the ViolationComparison shape via the API and set the value
                txt_Violation = sheet_dashboard.shapes['txtDViolationsDiff'].api
                txt_Violation.TextFrame2.TextRange.Text = f"${prev_ViolationComp} to ${lat_ViolationComp}"
                txt_Violation_diff = sheet_dashboard.shapes['txtViolationsDiff'].api
                txt_Violation_diff.TextFrame2.TextRange.Text = f"${diff_ViolationComp}"

                txt_CashCollected = sheet_dashboard.shapes['txtDCashCollectedDiff'].api
                txt_CashCollected.TextFrame2.TextRange.Text = f"${prev_CashCollectedComp} to ${lat_CashCollectedComp}"
                txt_CashCollected_diff = sheet_dashboard.shapes['txtCashCollectedDiff'].api
                txt_CashCollected_diff.TextFrame2.TextRange.Text = f"${diff_CashCollectedComp}"

                # # Access the OperatorChanges shape via the API and set the value
                # txt_Operator = sheet_dashboard.shapes['txtDOperatorsDiff'].api
                # txt_Operator.TextFrame2.TextRange.Text = f"{prev_OperatorChanges} to {lat_OperatorChanges} operators"
                # txt_Operator_diff = sheet_dashboard.shapes['txtOperatorsDiff'].api
                # txt_Operator_diff.TextFrame2.TextRange.Text = f"{diff_OperatorChanges} operators"
                
                # # Access the MissingDates shape via the API and set the value
                # txt_Date = sheet_dashboard.shapes['txtDDatesDiff'].api
                # txt_Date.TextFrame2.TextRange.Text = f"{prev_DateChanges} to {lat_DateChanges} days"
                # txt_Date_diff = sheet_dashboard.shapes['txtDatesDiff'].api
                # txt_Date_diff.TextFrame2.TextRange.Text = f"{diff_DateChanges} days"

            elif sheet_name == 'ADA':
                sheet_TotalRevHrsComparison = wb_comparison['TotalRevHrsComparison']
                lat_TotalRevHrsComp = f"{sum(cell.value for row in sheet_TotalRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_TotalRevHrsComp = f"{sum(cell.value for row in sheet_TotalRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_TotalRevHrsComp = f"{sum(cell.value for row in sheet_TotalRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_PRevHrsComparison = wb_comparison['%RevHrsComparison']
                lat_PRevHrs = f"{(lambda v: (sum(v)/len(v)*100 if v else 0))([cell.value for row in sheet_PRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None]):,.2f}"
                prev_PRevHrs = f"{(lambda v: (sum(v)/len(v)*100 if v else 0))([cell.value for row in sheet_PRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None]):,.2f}"
                diff_PRevHrs = f"{(lambda v: (sum(v)/len(v)*100 if v else 0))([cell.value for row in sheet_PRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None]):,.2f}"

                sheet_BonusHrsComparison = wb_comparison['BonusHrsComparison']
                lat_BonusHrs_Comp = f"{sum(cell.value for row in sheet_BonusHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,}"
                prev_BonusHrs_Comp = f"{sum(cell.value for row in sheet_BonusHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,}"
                diff_BonusHrs_Comp = f"{sum(cell.value for row in sheet_BonusHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,}"

                sheet_CoreRevComparison = wb_comparison['CoreRevComparison']
                lat_CoreRev_Comp = f"{sum(cell.value for row in sheet_CoreRevComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_CoreRev_Comp = f"{sum(cell.value for row in sheet_CoreRevComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_CoreRev_Comp = f"{sum(cell.value for row in sheet_CoreRevComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_TotEarningsComparison = wb_comparison['TotEarningsComparison']
                lat_TotEarnings_Comp = f"{sum(cell.value for row in sheet_TotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_TotEarnings_Comp = f"{sum(cell.value for row in sheet_TotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_TotEarnings_Comp = f"{sum(cell.value for row in sheet_TotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_DriversComparison = wb_comparison['DriversComparison']
                lat_Drivers_Comp = sum(1 for row in sheet_DriversComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None)
                prev_Drivers_Comp = sum(1 for row in sheet_DriversComparison.iter_rows(min_row=2, max_row=50, min_col=1, max_col=1) for cell in row if cell.value is not None)
                diff_Drivers_Comp = lat_Drivers_Comp - prev_Drivers_Comp


                sheet_dashboard = wb_dashboard.sheets[sheet_name]

                #For TotalRev Hours
                txt_TotalRevHrs = sheet_dashboard.shapes['txtDTotalRevHrsDiff'].api
                txt_TotalRevHrs.TextFrame2.TextRange.Text = f"${prev_TotalRevHrsComp} to ${lat_TotalRevHrsComp}"
                txt_TotalRevHrs_diff = sheet_dashboard.shapes['txtTotalRevHrsDiff'].api
                txt_TotalRevHrs_diff.TextFrame2.TextRange.Text = f"${diff_TotalRevHrsComp}"
            
                # For PRevHrs
                txt_PRevHrs = sheet_dashboard.shapes['txtDPRevHrsDiff'].api
                txt_PRevHrs.TextFrame2.TextRange.Text = f"{prev_PRevHrs}% to {lat_PRevHrs}%"
                txt_PRevHrs_diff = sheet_dashboard.shapes['txtPRevHrsDiff'].api
                txt_PRevHrs_diff.TextFrame2.TextRange.Text = f"{diff_PRevHrs}%"

                # For BonusHrs
                txt_BonusHrs = sheet_dashboard.shapes['txtDBonusHrsDiff'].api
                txt_BonusHrs.TextFrame2.TextRange.Text = f"{prev_BonusHrs_Comp} to {lat_BonusHrs_Comp} hours"
                txt_BonusHrs_diff = sheet_dashboard.shapes['txtBonusHrsDiff'].api
                txt_BonusHrs_diff.TextFrame2.TextRange.Text = f"{diff_BonusHrs_Comp} hours"

                # For CoreRev
                txt_CoreRev = sheet_dashboard.shapes['txtDCoreRevDiff'].api
                txt_CoreRev.TextFrame2.TextRange.Text = f"${prev_CoreRev_Comp} to ${lat_CoreRev_Comp}"
                txt_CoreRev_diff = sheet_dashboard.shapes['txtCoreRevDiff'].api
                txt_CoreRev_diff.TextFrame2.TextRange.Text = f"${diff_CoreRev_Comp}"

                # For TotEarnings
                txt_TotEarnings = sheet_dashboard.shapes['txtDTotEarningsDiff'].api
                txt_TotEarnings.TextFrame2.TextRange.Text = f"${prev_TotEarnings_Comp} to ${lat_TotEarnings_Comp}"
                txt_TotEarnings_diff = sheet_dashboard.shapes['txtTotEarningsDiff'].api
                txt_TotEarnings_diff.TextFrame2.TextRange.Text = f"${diff_TotEarnings_Comp}"

                # For Drivers
                txt_Drivers = sheet_dashboard.shapes['txtDDriversDiff'].api
                txt_Drivers.TextFrame2.TextRange.Text = f"{prev_Drivers_Comp} to {lat_Drivers_Comp} drivers"
                txt_Drivers_diff = sheet_dashboard.shapes['txtDriversDiff'].api
                txt_Drivers_diff.TextFrame2.TextRange.Text = f"{diff_Drivers_Comp} drivers"
            
            elif sheet_name == 'GOLINK':
                sheet_CoreHrsComparison = wb_comparison['CoreHoursComparison']
                lat_CoreHrsComp = f"{sum(cell.value for row in sheet_CoreHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_CoreHrsComp = f"{sum(cell.value for row in sheet_CoreHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_CoreHrsComp = f"{sum(cell.value for row in sheet_CoreHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_GPRevHrsComparison = wb_comparison['%RevHrsComparison']
                lat_GPRevHrsCompt = f"{(lambda v: (sum(v)/len(v)*100 if v else 0))([cell.value for row in sheet_GPRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None]):,.2f}"
                prev_GPRevHrsCompt = f"{(lambda v: (sum(v)/len(v)*100 if v else 0))([cell.value for row in sheet_GPRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None]):,.2f}"
                diff_GPRevHrsCompt = f"{(lambda v: (sum(v)/len(v)*100 if v else 0))([cell.value for row in sheet_GPRevHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None]):,.2f}"

                sheet_GTotEarningsComparison = wb_comparison['TotEarningsComparison']
                lat_GTotEarningsCompt = f"{sum(cell.value for row in sheet_GTotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_GTotEarningsCompt = f"{sum(cell.value for row in sheet_GTotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_GTotEarningsCompt = f"{sum(cell.value for row in sheet_GTotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_totDutyViolationsComparison = wb_comparison['TotDutyViolationComparison']
                lat_totDutyViolationsCompt = f"{sum(cell.value for row in sheet_totDutyViolationsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_totDutyViolationsCompt = f"{sum(cell.value for row in sheet_totDutyViolationsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_totDutyViolationsCompt = f"{sum(cell.value for row in sheet_totDutyViolationsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_GDriversComparison = wb_comparison['DriversComparison']
                lat_GDrivers_Comp = sum(1 for row in sheet_GDriversComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None)
                prev_GDrivers_Comp = sum(1 for row in sheet_GDriversComparison.iter_rows(min_row=2, max_row=50, min_col=1, max_col=1) for cell in row if cell.value is not None)
                diff_GDrivers_Comp = lat_GDrivers_Comp - prev_GDrivers_Comp

                sheet_dashboard = wb_dashboard.sheets[sheet_name]

                # For Core Hours
                txt_CoreHrs = sheet_dashboard.shapes['txtDCoreHrsDiff'].api
                txt_CoreHrs .TextFrame2.TextRange.Text = f"{prev_CoreHrsComp} to {lat_CoreHrsComp} hours"
                txt_CoreHrs_diff = sheet_dashboard.shapes['txtCoreHrsDiff'].api
                txt_CoreHrs_diff.TextFrame2.TextRange.Text = f"{diff_CoreHrsComp} hours"

                # For GPRevHrs
                txt_GPRevHrs = sheet_dashboard.shapes['txtDPRevHrsDiff'].api
                txt_GPRevHrs.TextFrame2.TextRange.Text = f"{prev_GPRevHrsCompt}% to {lat_GPRevHrsCompt}%"
                txt_GPRevHrs_diff = sheet_dashboard.shapes['txtPRevHrsDiff'].api
                txt_GPRevHrs_diff.TextFrame2.TextRange.Text = f"{diff_GPRevHrsCompt}%"

                # For GTotEarnings
                txt_GTotEarnings = sheet_dashboard.shapes['txtDTotEarningsDiff'].api
                txt_GTotEarnings.TextFrame2.TextRange.Text = f"${prev_GTotEarningsCompt} to ${lat_GTotEarningsCompt}"
                txt_GTotEarnings_diff = sheet_dashboard.shapes['txtTotEarningsDiff'].api
                txt_GTotEarnings_diff.TextFrame2.TextRange.Text = f"${diff_GTotEarningsCompt}"

                # For totDutyViolations
                txt_totDutyViolations = sheet_dashboard.shapes['txtDTotalViolationsDiff'].api
                txt_totDutyViolations.TextFrame2.TextRange.Text = f"${prev_totDutyViolationsCompt} to ${lat_totDutyViolationsCompt}"
                txt_totDutyViolations_diff = sheet_dashboard.shapes['txtTotalViolationsDiff'].api
                txt_totDutyViolations_diff.TextFrame2.TextRange.Text = f"${diff_totDutyViolationsCompt}"

                # For Drivers
                txt_GDrivers = sheet_dashboard.shapes['txtDDriversDiff'].api
                txt_GDrivers.TextFrame2.TextRange.Text = f"{prev_GDrivers_Comp} to {lat_GDrivers_Comp} drivers"
                txt_GDrivers_diff = sheet_dashboard.shapes['txtDriversDiff'].api
                txt_GDrivers_diff.TextFrame2.TextRange.Text = f"{diff_GDrivers_Comp} drivers"

            elif sheet_name == 'ADAGOLINK':
                sheet_StdByPayHrsComparison = wb_comparison['StdByPayHrsComparison']
                lat_StdByPayHrsComp = f"{sum(cell.value for row in sheet_StdByPayHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_StdByPayHrsComp = f"{sum(cell.value for row in sheet_StdByPayHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_StdByPayHrsComp = f"{sum(cell.value for row in sheet_StdByPayHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_StByExtraHrsComparison = wb_comparison['StByExtraHrsComparison']
                lat_StdByExtraHrsComp = f"{sum(cell.value for row in sheet_StByExtraHrsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_StdByExtraHrsComp = f"{sum(cell.value for row in sheet_StByExtraHrsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_StdByExtraHrsComp = f"{sum(cell.value for row in sheet_StByExtraHrsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_StdByTotEarningsComparison = wb_comparison['TotEarningsComparison']
                lat_StdByTotEarningsComp = f"{sum(cell.value for row in sheet_StdByTotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
                prev_StdByTotEarningsComp = f"{sum(cell.value for row in sheet_StdByTotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
                diff_StdByTotEarningsComp = f"{sum(cell.value for row in sheet_StdByTotEarningsComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

                sheet_dashboard = wb_dashboard.sheets[sheet_name]

                # For StdByPayHrs
                txt_StdByPayHrs = sheet_dashboard.shapes['txtDstdPayHrsDiff'].api
                txt_StdByPayHrs.TextFrame2.TextRange.Text = f"${prev_StdByPayHrsComp} to ${lat_StdByPayHrsComp}"
                txt_StdByPayHrs_diff = sheet_dashboard.shapes['txtstdPayHrsDiff'].api
                txt_StdByPayHrs_diff.TextFrame2.TextRange.Text = f"${diff_StdByPayHrsComp}"

                # For StdByExtraHrs
                txt_StdByExtraHrs = sheet_dashboard.shapes['txtDstdExtraHrsDiff'].api
                txt_StdByExtraHrs.TextFrame2.TextRange.Text = f"{prev_StdByExtraHrsComp} to {lat_StdByExtraHrsComp} hours"
                txt_StdByExtraHrs_diff = sheet_dashboard.shapes['txtstdExtraHrsDiff'].api
                txt_StdByExtraHrs_diff.TextFrame2.TextRange.Text = f"{diff_StdByExtraHrsComp} hours"

                # For StdByTotEarnings
                txt_StdByTotEarnings = sheet_dashboard.shapes['txtDTotEarningsDiff'].api
                txt_StdByTotEarnings.TextFrame2.TextRange.Text = f"${prev_StdByTotEarningsComp} to ${lat_StdByTotEarningsComp}"
                txt_StdByTotEarnings_diff = sheet_dashboard.shapes['txtTotEarningsDiff'].api
                txt_StdByTotEarnings_diff.TextFrame2.TextRange.Text = f"${diff_StdByTotEarningsComp}"

            # paste_picture(comparison_files, dashboard_file)

            
            if sheet_name == 'Dashboard':
                # Run the VBA macro to update the color based on the status
                try:
                    # Parameters: TextBox name and status
                    textBoxName = "TextBox 90"
                    status = full_amount_diff_status  # Use the status from the comparison file

                    # Call the VBA macro to update color
                    wb_dashboard.macro("UpdateTextBoxColor")(sheet_name, textBoxName, status)
                    logger.info(f"Successfully updated color for {textBoxName} with status '{status}'.")
                except Exception as e:
                    logger.error(f"An error occurred: {e}")

            # Run the VBA macro to update the color based on the values
            try:
                # Parameters: TextBox names and corresponding values
                # textBoxNames = ["txtHTotalRevDiff", "txtLLeaseDiff", "txtViolationsDiff", "txtOperatorsDiff"]
                # values = [diff_HTotalRev, diff_LeaseComp, diff_ViolationComp, diff_OperatorChanges]

                if sheet_name == 'Dashboard':
                    textBoxNames = ["txtHTTLRevtDiff", "txtLLeaseDiff", "txtViolationsDiff", "txtCashCollectedDiff"]
                    values = [diff_HTTLRevComp, diff_LeaseComp, diff_ViolationComp, diff_CashCollectedComp]
                elif sheet_name == 'ADA':
                    textBoxNames = ["txtTotalRevHrsDiff", "txtPRevHrsDiff", "txtBonusHrsDiff", "txtCoreRevDiff", "txtTotEarningsDiff", "txtDriversDiff"]
                    values = [diff_TotalRevHrsComp, diff_PRevHrs, diff_BonusHrs_Comp, diff_CoreRev_Comp, diff_TotEarnings_Comp, diff_Drivers_Comp]
                elif sheet_name == 'GOLINK':
                    textBoxNames = ["txtCoreHrsDiff", "txtPRevHrsDiff", "txtTotEarningsDiff", "txtTotalViolationsDiff", "txtDriversDiff"]
                    values = [diff_CoreHrsComp, diff_GPRevHrsCompt, diff_GTotEarningsCompt, diff_totDutyViolationsCompt, diff_GDrivers_Comp]
                elif sheet_name == 'ADAGOLINK':
                    textBoxNames = ["txtstdPayHrsDiff", "txtstdExtraHrsDiff", "txtTotEarningsDiff"]
                    values = [diff_StdByPayHrsComp, diff_StdByExtraHrsComp, diff_StdByTotEarningsComp]

                # Loop through the text boxes and update colors based on the values
                for i, textBoxName in enumerate(textBoxNames):
                    wb_dashboard.macro("UpdateSummaryColor")(sheet_name, textBoxName, values[i])
                    logger.info(f"Successfully updated color for {textBoxName} with value '{values[i]}'.")
            except Exception as e:
                logger.error(f"A TextBox error occurred: {e}")

        # Save the changes to the dashboard workbook
        wb_dashboard.save()
        wb_dashboard.close()
        logger.info(f"{dashboard_file} has been successfully updated and saved.")

        # paste_picture(comparison_files, dashboard_file)
        app.quit()
        time.sleep(2)
        paste_picture()
        time.sleep(2)

    except Exception as e:
        logger.info(f"A Dashboard error occurred: {e}")
    # finally:
    #     wb_dashboard.save()
    #     wb_dashboard.close()
        # app.quit()
        
        # Reopen the Excel file

        # app = xw.App(visible=True)  # Open Excel with the app visible
        # wb_dashboard = app.books.open(dashboard_file)  # Reopen the file

def paste_picture():
    comparison_files = [
        ('ComparedResults/DIV3_Main_Tables.xlsx', 'Dashboard'),
        ('ComparedResults/DIV3_ADA_Tables.xlsx', 'ADA'),
        ('ComparedResults/DIV3_GOLINK_Tables.xlsx', 'GOLINK'),
        ('ComparedResults/DIV3_ADAGOLINK_Tables.xlsx', 'ADAGOLINK')
    ]
    
    # Target cells for each sheet in the comparison file
    target_cells = {
        'HourlyTTLRevComparison': (11, 12),
        'LiftLeaseComparison': (11, 22),
        'ViolationsComparison': (11, 30),
        'CashCollectedComparison': (11, 41)
    }

    relative_dashboard_path = "ComparedResults\\Dashboard.xlsm"
    # Get the absolute path of the current script's directory
    # script_dir = os.path.dirname(os.path.realpath(__file__))

    # # Build the full path to the dashboard file by joining the script directory and the relative path
    # dashboard_file = os.path.join(script_dir, relative_dashboard_path)

    # Get the base directory correctly whether running as script or PyInstaller executable
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)  # Executable folder
    else:
        script_dir = os.path.dirname(os.path.realpath(__file__))  # Script folder

    dashboard_file = os.path.join(script_dir, "ComparedResults", "Dashboard.xlsm")

    excel = None

    try:
        # Initialize Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True  # Set to True for debugging

        # Check if the file exists
        if not os.path.exists(dashboard_file):
            logger.info(f"Error: Dashboard file does not exist at {dashboard_file}")
            return
        
        # Open the Dashboard workbook
        wb_dashboard = excel.Workbooks.Open(dashboard_file)
        if wb_dashboard is None:
            logger.info(f"Failed to open the Dashboard workbook at {dashboard_file}")
            return

        # Delete existing pictures if they exist
        sheet_names = ['Dashboard', 'ADA']
        for target_sheet_name in sheet_names:
            ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
            ws_dashboard.Activate()
            # for picture_name in ['ViolationsTable', 'HoursTable', 'OperatorTable', 'LeaseTable']:
            picname = ['HTTLRevTable', 'LeaseTable', 'ViolationsTable', 'CashTable']
            if target_sheet_name == 'ADA': 
                picname = ['TotalRevHrsTable', 'PRevHrsTable', 'BonusHrsTable', 'CoreRevTable', 'TotEarningsTable', 'DriversTable']
            elif target_sheet_name == 'GOLINK':
                picname = ['TotalCoreHrsTable', 'PRevHrsTable', 'TotEarningsTable', 'TotalViolationsTable', 'DriversTable']
            elif target_sheet_name == 'ADAGOLINK':
                picname = ['stdPayHrsTable', 'stdExtraHrsTable', 'TotEarningsTable']
            for picture_name in picname:
                logger.info(f"Picture name: {picture_name} found in {target_sheet_name}") 
                try:
                    ws_dashboard.Shapes(picture_name).Delete()  # Attempt to delete the picture
                    logger.info(f"Deleted existing picture: {picture_name} in {target_sheet_name}")
                except Exception:
                    logger.info(f"No existing picture named {picture_name} found in {target_sheet_name}")  

        time.sleep(2)
        
        # Process each comparison file
        for comparison_file, target_sheet_name in comparison_files:
            # Build the full path for the comparison file
            comparison_file_path = os.path.join(script_dir, comparison_file)

            target_cells = {
                'HourlyTTLRevComparison': (11, 12),
                'LiftLeaseComparison': (11, 22),
                'ViolationsComparison': (11, 30),
                'CashCollectedComparison': (11, 41)
            }
                    
            # Check if the comparison file exists
            if not os.path.exists(comparison_file_path):
                logger.info(f"Error: Comparison file does not exist at {comparison_file_path}")
                continue

            # Open the comparison workbook
            wb_comparison = excel.Workbooks.Open(comparison_file_path)
            if wb_comparison is None:
                logger.info(f"Failed to open the comparison workbook at {comparison_file_path}")
                continue
            if target_sheet_name == 'ADA':
                target_cells = {
                    'TotalRevHrsComparison': (5, 12),
                    '%RevHrsComparison': (5, 20),
                    'BonusHrsComparison': (5, 28),
                    'CoreRevComparison': (37, 12),
                    'TotEarningsComparison': (37, 20),
                    'DriversComparison': (37, 28)
                }
            elif target_sheet_name == 'GOLINK':
                target_cells = {
                    'CoreHoursComparison': (10, 12),
                    '%RevHrsComparison': (10, 20),
                    'TotEarningsComparison': (23, 12),
                    'TotDutyViolationComparison': (23, 20),
                    'DriversComparison': (10, 28)
                }
            elif target_sheet_name == 'ADAGOLINK':
                target_cells = {
                    'StdByPayHrsComparison': (10, 12),
                    'StByExtraHrsComparison': (10, 20),
                    'TotEarningsComparison': (10, 28)
                }


            # Process each sheet in the comparison file (ViolationComparison, HTotalRevComparison, OperatorChanges)
            for sheet_name, target_cell in target_cells.items():
                sheet = wb_comparison.Sheets(sheet_name)
                table_width = 0
                table_height = 0
                table_name = None
                if sheet is None:
                    logger.info(f"Failed to access the '{sheet_name}' sheet in {comparison_file_path}")
                    continue

                # Get the used range
                used_range = sheet.UsedRange

                time.sleep(1)
                if used_range.Rows.Count > 0 and used_range.Columns.Count > 0:
                    try:
                        table_width = (used_range.Width) # Get the width of the used range
                        table_height = (used_range.Height) # Get the height of the used range
                        used_range.CopyPicture(Format=2)
                        logger.info(f"Copied picture from {sheet_name}, width: {table_width*0.0352778:.2f} cm, height: {table_height*0.0352778:.2f} cm")
                    except Exception as e:
                        logger.error(f"Failed to copy picture from {sheet_name}: {e}")
                        continue
                else:
                    logger.error(f"Skipping {sheet_name}: No data in the range.")
                    continue  # Skip further processing for this sheet

                # Activate the target sheet in the Dashboard workbook
                ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
                if ws_dashboard is None:
                    logger.info(f"Failed to access the sheet '{target_sheet_name}' in the Dashboard workbook.")
                    wb_comparison.Close(SaveChanges=False)
                    continue
                
                time.sleep(1)
                # Paste as a picture into the target sheet
                ws_dashboard.Activate()
                row, col = target_cell
                target_cell_range = ws_dashboard.Cells(row, col)  # Adjust as needed
                ws_dashboard.Paste(target_cell_range)

                # Position and resize the pasted picture
                pasted_picture = ws_dashboard.Shapes(ws_dashboard.Shapes.Count)
                pasted_picture.Left = target_cell_range.Left
                pasted_picture.Top = target_cell_range.Top

                # Name the pasted picture according to the sheet
                if sheet_name == 'LiftLeaseComparison':
                    pasted_picture.Name = 'LeaseTable'
                elif sheet_name == 'ViolationsComparison':
                    pasted_picture.Name = 'ViolationsTable'
                elif sheet_name == 'HourlyTTLRevComparison':
                    pasted_picture.Name = 'HTTLRevTable'
                elif sheet_name == 'CashCollectedComparison':
                    pasted_picture.Name = 'CashTable'
                elif sheet_name == 'TotalRevHrsComparison':
                    pasted_picture.Name = 'TotalRevHrsTable'
                elif sheet_name == '%RevHrsComparison':
                    pasted_picture.Name = 'PRevHrsTable'
                elif sheet_name == 'BonusHrsComparison':
                    pasted_picture.Name = 'BonusHrsTable'
                elif sheet_name == 'CoreRevComparison':
                    pasted_picture.Name = 'CoreRevTable'
                elif sheet_name == 'TotEarningsComparison':
                    pasted_picture.Name = 'TotEarningsTable'
                elif sheet_name == 'DriversComparison':
                    pasted_picture.Name = 'DriversTable'
                elif sheet_name == 'CoreHoursComparison':
                    pasted_picture.Name = 'TotalCoreHrsTable'
                elif sheet_name == 'TotDutyViolationComparison':
                    pasted_picture.Name = 'TotalViolationsTable'
                elif sheet_name == 'StdByPayHrsComparison':
                    pasted_picture.Name = 'stdPayHrsTable'
                elif sheet_name == 'StByExtraHrsComparison':
                    pasted_picture.Name = 'stdExtraHrsTable'

                table_name = pasted_picture.Name
                logger.info(f"Table Name: {table_name}")

                time.sleep(1)
                logger.info(f"'{pasted_picture.Name}' successfully pasted in the {sheet_name} Sheet.")

                # Adjust the container size based on the table size
                # Adjust only the container that matches the table name
                if table_name:
                    container_name = table_name.replace("Table", "Container")  # Match the container name
                    try:
                        container = ws_dashboard.Shapes(container_name)
                        # container.Width = table_width + 95  # Add 3.35 cm to width
                        container.Width = table_width + 2 # Add 3.35 cm to width
                        # container.Height = table_height + 123  # Add 4.33 cm to height
                        container.Height = table_height + 56  # Add 4.33 cm to height
                        logger.info(f"Resized {container_name} to width: {(container.Width)*0.0352778:.2f} cm, height: {(container.Height)*0.0352778:.2f} cm")
                    except Exception as e:
                        logger.error(f"Failed to resize {container_name}: {e}")

                # wb_comparison.Save()
                # wb_dashboard.Save()
        wb_comparison.Save()
        wb_comparison.Close() 
            # Close the comparison workbook without saving
            # wb_comparison.Close(SaveChanges=True)

        # Save and close the Dashboard workbook
        wb_dashboard.Save()
        wb_dashboard.Close()
        excel.Quit()

        logger.info("Data pasted as pictures successfully.")

        app = xw.App(visible=True)  # Open Excel with the app visible
        wb_dashboard = app.books.open(dashboard_file)  # Reopen the file
    except Exception as e:
        logger.error(f"An error occurred: {e}")

        if 'wb_dashboard' in locals() and wb_dashboard:
            wb_dashboard.Close(SaveChanges=False)

        if excel:
            excel.Quit()
            del excel
    finally:
        # Ensure Excel is properly quit and the object is released
        if excel:
            excel.Quit()  # Close the Excel application
            del excel 

# def adjust_container(table_name, table_width, table_height):
#     for target_sheet_name in ['Dashboard', 'CCCTA', 'LAVTA']:
#             ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
#             ws_dashboard.Activate()
#             for picture_name in ['TripsTable', 'HoursTable', 'OperatorTable', 'LeaseTable']:
#                 try:
#                     ws_dashboard.Shapes(picture_name).Delete()  # Attempt to delete the picture
#                     logger.info(f"Deleted existing picture: {picture_name} in {target_sheet_name}")
#                 except Exception:
#                     logger.info(f"No existing picture named {picture_name} found in {target_sheet_name}")  

#     # Get the active table and container dimensions
#     if table_name == 'TripsTable':
#     elif table_name == 'HoursTable':
#     elif table_name == 'OperatorTable':
#     elif table_name == 'LeaseTable':

#     # Calculate the aspect ratios
#     table_aspect_ratio = table_width / table_height
#     container_aspect_ratio = container_width / container_height

#     # Determine the scaling factor
#     if table_aspect_ratio > container_aspect_ratio:
#         # Scale based on width
#         scale_factor = container_width / table_width
#     else:
#         # Scale based on height
#         scale_factor = container_height / table_height

#     # Calculate the new dimensions
#     new_width = table_width * scale_factor
#     new_height = table_height * scale_factor

#     return new_width, new_height

# if __name__ == '__main__':
    
#     comparison_files = [
#         ('Compared Results/Full_Comparison.xlsx', 'Dashboard'),
#         ('Compared Results/CCCTA_Comparison.xlsx', 'CCCTA'),
#         ('Compared Results/LAVTA_Comparison.xlsx', 'LAVTA')
#     ]
#     dashboard_file = 'Compared Results/Dashboard.xlsm'
#     main(file_previous, file_latest)
#     paste_picture(comparison_files, dashboard_file)
#     paste_picture(comparison_files, dashboard_file)
