import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import filedialog
import dashboard as db
import time

log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
# Set up logging
logging.basicConfig(
    level=logging.DEBUG, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('Logs/Comparison.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Declare global variables
file_entry_previous = None
file_entry_latest = None

prev_operator, prev_date = 0, 0
lat_operator, lat_date = 0, 0

calculated_amount = []
# calculated_totals = 0

def load_sheets(file_previous, file_latest):
    try:
        logger.info("Loading sheets from the provided Excel files.")

        # Define possible sheet names
        sheet_options = {
            "PR": ["PR"],
            "ADA_Hours": ["ADAHours"],
            "ADA": ["ADA"],
            "GOLINK_Hours": ["GOLINKHours"],
            "GOLINK": ["GOLINK"],
            "StandbyADAGOLINK": ["StandbyADAGOLINK"],
            "Operator": ["Div3PartnerList"],
            "Deductions": ["Deductions&OtherEarnings"]
        }
        
        def load_sheet(file, sheet_names):
            """Tries to load the first available sheet from a list of sheet names."""
            available_sheets = pd.ExcelFile(file).sheet_names
            for sheet in sheet_names:
                if sheet in available_sheets:
                    return pd.read_excel(file, sheet_name=sheet)
            logger.warning(f"None of the specified sheets {sheet_names} found in {file}")
            return None  # Return None if no matching sheet is found
        
        sheet_pr_previous = load_sheet(file_previous, sheet_options["PR"])
        sheet_pr_latest = load_sheet(file_latest, sheet_options["PR"])

        sheet_ADA_Hours_previous = load_sheet(file_previous, sheet_options["ADA_Hours"])
        sheet_ADA_Hours_latest = load_sheet(file_latest, sheet_options["ADA_Hours"])

        sheet_ADA_previous = load_sheet(file_previous, sheet_options["ADA"])
        sheet_ADA_latest = load_sheet(file_latest, sheet_options["ADA"])

        sheet_GOLINK_Hours_previous = load_sheet(file_previous, sheet_options["GOLINK_Hours"])
        sheet_GOLINK_Hours_latest = load_sheet(file_latest, sheet_options["GOLINK_Hours"])

        sheet_GOLINK_previous = load_sheet(file_previous, sheet_options["GOLINK"])
        sheet_GOLINK_latest = load_sheet(file_latest, sheet_options["GOLINK"])

        sheet_StandbyADAGOLINK_previous = load_sheet(file_previous, sheet_options["StandbyADAGOLINK"])
        sheet_StandbyADAGOLINK_latest = load_sheet(file_latest, sheet_options["StandbyADAGOLINK"])

        sheet_operator_previous = load_sheet(file_previous, sheet_options["Operator"])
        sheet_operator_latest = load_sheet(file_latest, sheet_options["Operator"])

        sheet_deductions_previous = load_sheet(file_previous, sheet_options["Deductions"])
        sheet_deductions_latest = load_sheet(file_latest, sheet_options["Deductions"])
        
        logger.info("Sheets loaded successfully.")
        return sheet_pr_previous, sheet_pr_latest, sheet_ADA_previous, sheet_ADA_latest, sheet_ADA_Hours_previous, sheet_ADA_Hours_latest, sheet_GOLINK_Hours_previous, sheet_GOLINK_Hours_latest, sheet_GOLINK_previous, sheet_GOLINK_latest, sheet_StandbyADAGOLINK_previous, sheet_StandbyADAGOLINK_latest, sheet_operator_previous, sheet_operator_latest, sheet_deductions_previous, sheet_deductions_latest
    except FileNotFoundError as e:
        logger.error(f"File not found: {e}")
        raise
    except Exception as e:
        logger.error(f"Error loading sheets: {e}")
        raise

def clean_currency(value):
    try:
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
            return round(float(value), 2) if value else None
        return value
    except ValueError:
        logger.error(f"Error cleaning currency value: {value}")
        return None

def calculate_totals(deductions_sheet, pr_sheet):
    try:
        if not isinstance(deductions_sheet, pd.DataFrame):
            raise TypeError("deductions_sheet should be a pandas DataFrame.")
        if not isinstance(pr_sheet, pd.DataFrame):
            raise TypeError("pr_sheet should be a pandas DataFrame.")

        calculated_totals = 0

        client_header_index = 3  # Manually setting starting point

        # Find the first empty row after the client header row
        empty_row_index = pr_sheet.iloc[client_header_index + 1:, 0].isna().idxmax()

        if pd.isna(empty_row_index):
            next_header_index = pr_sheet.shape[0]
        else:
            next_header_index = empty_row_index + client_header_index + 1

        # Extract the partner rows
        partner_rows = pr_sheet.iloc[client_header_index + 1:next_header_index]

        # Filter out rows where the Partner (first column) is empty or 0
        valid_partner_rows = partner_rows[
            partner_rows.iloc[:, 0].notna() & (partner_rows.iloc[:, 0] != 0)
        ]

        # Get partner names as a list
        partner_names = valid_partner_rows.iloc[:, 0].astype(str).str.strip().tolist()

        logger.info(f"Total valid partners: {len(valid_partner_rows)}")
        logger.info(f"Partner names: {partner_names}")

        total_amount = 0

        # Sum the Amount column (column 13, index 13) for valid partners
        total_amount = valid_partner_rows.iloc[:, 14].sum()

        calculated_totals += total_amount

        logger.info(f"Total amount: {total_amount}")
        logger.info(f"Total calculated amount: {calculated_totals}")

        return calculated_totals

    except Exception as e:
        logger.error(f"Error calculating totals: {e}")
        raise

def compare_totals(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Totals between previous and latest values.")

        # Ensure inputs are numeric
        if not isinstance(sheet_previous, (int, float)) or not isinstance(sheet_latest, (int, float)):
            raise TypeError("Both sheet_previous and sheet_latest must be numeric values.")
        
        # Create a DataFrame to store results
        deductions_comparison = pd.DataFrame({
            "LATEST": [sheet_latest],
            "PREVIOUS": [sheet_previous],
            "DIFFERENCE": [sheet_latest - sheet_previous]
        }).round(2)

        # Add CHANGE column based on the difference
        deductions_comparison["CHANGE"] = deductions_comparison["DIFFERENCE"].apply(
            lambda diff: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
        )

        logger.info("Totals comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing totals: {e}")
        raise

def compare_TTL_Rev(sheet_previous, sheet_latest):
    try:
        if not isinstance(sheet_previous, pd.DataFrame):
            raise TypeError("sheet_previous should be a pandas DataFrame.")
        if not isinstance(sheet_latest, pd.DataFrame):
            raise TypeError("sheet_latest should be a pandas DataFrame.")
        
        sheet_previous = sheet_previous.dropna(subset=[sheet_previous.columns[0], sheet_previous.columns[3]])
        sheet_latest = sheet_latest.dropna(subset=[sheet_latest.columns[0], sheet_latest.columns[3]])

        previous_data = sheet_previous.iloc[:, [0, 3]].copy()
        latest_data = sheet_latest.iloc[:, [0, 3]].copy()

        previous_data.columns = ["PARTNER", "PREVIOUS"]
        latest_data.columns = ["PARTNER", "LATEST"]

        comparison = pd.merge(latest_data, previous_data, on="PARTNER", how="outer")

        comparison["LATEST"] = comparison["LATEST"].fillna(0)
        comparison["PREVIOUS"] = comparison["PREVIOUS"].fillna(0)

        # Convert columns to numeric
        comparison["LATEST"] = pd.to_numeric(comparison["LATEST"], errors='coerce').fillna(0)
        comparison["PREVIOUS"] = pd.to_numeric(comparison["PREVIOUS"], errors='coerce').fillna(0)

        comparison["CHANGE"] = comparison["LATEST"] - comparison["PREVIOUS"]

        comparison = comparison[["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]]

        comparison = comparison.drop(index=0)

        return comparison

    except Exception as e:
        logger.error(f"Error comparing TTL Rev: {e}")
        raise

def compare_liftlease(sheet_previous, sheet_latest):
    try:
        if not isinstance(sheet_previous, pd.DataFrame):
            raise TypeError("sheet_previous should be a pandas DataFrame.")
        if not isinstance(sheet_latest, pd.DataFrame):
            raise TypeError("sheet_latest should be a pandas DataFrame.")
        
        sheet_previous = sheet_previous.dropna(subset=[sheet_previous.columns[0], sheet_previous.columns[9]])
        sheet_latest = sheet_latest.dropna(subset=[sheet_latest.columns[0], sheet_latest.columns[9]])

        previous_data = sheet_previous.iloc[:, [0, 9]].copy()
        latest_data = sheet_latest.iloc[:, [0, 9]].copy()

        previous_data.columns = ["PARTNER", "PREVIOUS"]
        latest_data.columns = ["PARTNER", "LATEST"]

        comparison = pd.merge(latest_data, previous_data, on="PARTNER", how="outer")

        comparison["LATEST"] = comparison["LATEST"].fillna(0)
        comparison["PREVIOUS"] = comparison["PREVIOUS"].fillna(0)

        # Convert columns to numeric
        comparison["LATEST"] = pd.to_numeric(comparison["LATEST"], errors='coerce').fillna(0)
        comparison["PREVIOUS"] = pd.to_numeric(comparison["PREVIOUS"], errors='coerce').fillna(0)

        comparison["CHANGE"] = comparison["LATEST"] - comparison["PREVIOUS"]

        comparison = comparison[["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]]

        comparison = comparison.drop(index=0)

        return comparison

    except Exception as e:
        logger.error(f"Error comparing Lift Lease: {e}")
        raise

def compare_violations(sheet_previous, sheet_latest):
    try:
        if not isinstance(sheet_previous, pd.DataFrame):
            raise TypeError("sheet_previous should be a pandas DataFrame.")
        if not isinstance(sheet_latest, pd.DataFrame):
            raise TypeError("sheet_latest should be a pandas DataFrame.")
        
        sheet_previous = sheet_previous.dropna(subset=[sheet_previous.columns[0], sheet_previous.columns[10]])
        sheet_latest = sheet_latest.dropna(subset=[sheet_latest.columns[0], sheet_latest.columns[10]])

        previous_data = sheet_previous.iloc[:, [0, 10]].copy()
        latest_data = sheet_latest.iloc[:, [0, 10]].copy()

        previous_data.columns = ["PARTNER", "PREVIOUS"]
        latest_data.columns = ["PARTNER", "LATEST"]

        comparison = pd.merge(latest_data, previous_data, on="PARTNER", how="outer")

        comparison["LATEST"] = comparison["LATEST"].fillna(0)
        comparison["PREVIOUS"] = comparison["PREVIOUS"].fillna(0)

        # Convert columns to numeric
        comparison["LATEST"] = pd.to_numeric(comparison["LATEST"], errors='coerce').fillna(0)
        comparison["PREVIOUS"] = pd.to_numeric(comparison["PREVIOUS"], errors='coerce').fillna(0)

        comparison["CHANGE"] = comparison["LATEST"] - comparison["PREVIOUS"]

        comparison = comparison[["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]]

        comparison = comparison.drop(index=0)

        return comparison

    except Exception as e:
        logger.error(f"Error comparing Violations: {e}")
        raise

def compare_cash_collected(sheet_previous, sheet_latest):
    try:
        if not isinstance(sheet_previous, pd.DataFrame):
            raise TypeError("sheet_previous should be a pandas DataFrame.")
        if not isinstance(sheet_latest, pd.DataFrame):
            raise TypeError("sheet_latest should be a pandas DataFrame.")
        
        sheet_previous = sheet_previous.dropna(subset=[sheet_previous.columns[0], sheet_previous.columns[11]])
        sheet_latest = sheet_latest.dropna(subset=[sheet_latest.columns[0], sheet_latest.columns[11]])

        previous_data = sheet_previous.iloc[:, [0, 11]].copy()
        latest_data = sheet_latest.iloc[:, [0, 11]].copy()

        previous_data.columns = ["PARTNER", "PREVIOUS"]
        latest_data.columns = ["PARTNER", "LATEST"]

        comparison = pd.merge(latest_data, previous_data, on="PARTNER", how="outer")

        comparison["LATEST"] = comparison["LATEST"].fillna(0)
        comparison["PREVIOUS"] = comparison["PREVIOUS"].fillna(0)

        # Convert columns to numeric
        comparison["LATEST"] = pd.to_numeric(comparison["LATEST"], errors='coerce').fillna(0)
        comparison["PREVIOUS"] = pd.to_numeric(comparison["PREVIOUS"], errors='coerce').fillna(0)

        comparison["CHANGE"] = comparison["LATEST"] - comparison["PREVIOUS"]

        comparison = comparison[["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]]

        comparison = comparison.drop(index=0)

        return comparison

    except Exception as e:
        logger.error(f"Error comparing Cash Collected: {e}")
        raise

def compare_TotRevHrs(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Total Revenue Hours")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Total Revenue Hours"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Total Revenue Hours"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["Total Revenue Hours_LATEST"] - comparison["Total Revenue Hours_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Total Revenue Hours_LATEST", "Total Revenue Hours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of Total Revenue Hours and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Total Revenue Hours: {e}")
        raise

def compare_PRevHrs(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing % of Revenue Hours to Forecast")

        # Group by PARTNER NAME and take the average of each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["% of Revenue Hours to Forecast"].mean()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["% of Revenue Hours to Forecast"].mean()


        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["% of Revenue Hours to Forecast_LATEST"] - comparison["% of Revenue Hours to Forecast_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["% of Revenue Hours to Forecast_LATEST", "% of Revenue Hours to Forecast_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of % of Revenue Hours to Forecast and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing % of Revenue Hours to Forecast: {e}")
        raise

def compare_BonusHrs(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Bonus Hours")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Bonus Hours"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Bonus Hours"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["Bonus Hours_LATEST"] - comparison["Bonus Hours_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Bonus Hours_LATEST", "Bonus Hours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of Bonus Hours and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Bonus Hours: {e}")
        raise

def compare_CoreRev(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Core Revenue")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Core Revenue"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Core Revenue"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["Core Revenue_LATEST"] - comparison["Core Revenue_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Core Revenue_LATEST", "Core Revenue_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of Core Revenue and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Core Revenue: {e}")
        raise

def compare_TotalEarnings(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Total Earnings")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Total Earnings"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Total Earnings"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["Total Earnings_LATEST"] - comparison["Total Earnings_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Total Earnings_LATEST", "Total Earnings_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of Total Earnings and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Total Earnings: {e}")
        raise

def compare_drivers(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing OS Operators between previous and latest sheets.")

        # Extract unique Drivers and sort them
        previous_values = sheet_previous[["Driver Name"]].drop_duplicates().rename(columns={"Driver Name": "PREVIOUS"}).sort_values(by="PREVIOUS")
        latest_values = sheet_latest[["Driver Name"]].drop_duplicates().rename(columns={"Driver Name": "LATEST"}).sort_values(by="LATEST")

        # Reset index to align properly
        previous_values = previous_values.reset_index(drop=True)
        latest_values = latest_values.reset_index(drop=True)

        # Ensure both columns have the same length by padding with empty strings
        max_length = max(len(previous_values), len(latest_values))
        previous_values = previous_values.reindex(range(max_length), fill_value="")
        latest_values = latest_values.reindex(range(max_length), fill_value="")

        # Combine into a single DataFrame
        comparison = pd.concat([previous_values, latest_values], axis=1)

        logger.info("Drivers comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Drivers: {e}")
        raise

def compare_CoreHrs(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Core Hours Worked")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Core Hours Worked"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Core Hours Worked"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["Core Hours Worked_LATEST"] - comparison["Core Hours Worked_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Core Hours Worked_LATEST", "Core Hours Worked_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of Core Hours Worked and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Core Hours Worked: {e}")
        raise

def compare_TotDutyViolations(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Total Duty Violation")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Total Duty Violation"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Total Duty Violation"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["Total Duty Violation_LATEST"] - comparison["Total Duty Violation_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Total Duty Violation_LATEST", "Total Duty Violation_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of Total Duty Violation and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Total Duty Violation: {e}")
        raise

def compare_stdPayHrs(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing StandBy Pay Hours")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["StandBy Pay Hours"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["StandBy Pay Hours"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["StandBy Pay Hours_LATEST"] - comparison["StandBy Pay Hours_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["StandBy Pay Hours_LATEST", "StandBy Pay Hours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of StandBy Pay Hours and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing StandBy Pay Hours: {e}")
        raise

def comapre_stdExtraHrs(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing StandbyExtraHours")

        # Group by PARTNER NAME and sum each metric
        prev_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["StandbyExtraHours"].sum()
        latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["StandbyExtraHours"].sum()

        # Merge the grouped dataframes
        comparison = latest_grouped.merge(
            prev_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

       # Calculate the change
        comparison["CHANGE"] = comparison["StandbyExtraHours_LATEST"] - comparison["StandbyExtraHours_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["StandbyExtraHours_LATEST", "StandbyExtraHours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Comparison of StandbyExtraHours and other metrics completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing StandbyExtraHours: {e}")
        raise

def apply_formatting(sheet_name, wb):
    try:
        logger.info(f"Applying formatting to sheet: {sheet_name}.")
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2.3

        # for col in ws.columns:
        #     max_length = 0
        #     column = col[0].column_letter
        #     for cell in col:
        #         try:
        #             if cell.value:
        #                 cell_length = len(str(cell.value))
        #                 if cell_length > max_length:
        #                     max_length = cell_length
        #         except:
        #             pass
        #     if ws[column + '1'].value:
        #         header_length = len(str(ws[column + '1'].value))
        #         if header_length > max_length:
        #             max_length = header_length

        #     adjusted_width = (max_length) * 1.2 # add scaling
        #     ws.column_dimensions[column].width = adjusted_width

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if ws[1][cell.column - 1].value.lower() == "change":
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                    elif isinstance(cell.value, str):
                        if cell.value.lower() in ["increased", "added"]:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value.lower() in ["decreased", "removed"]:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
        logger.info(f"Formatting applied successfully to sheet: {sheet_name}.")
    except Exception as e:
        logger.error(f"Error applying formatting to sheet {sheet_name}: {e}")
        raise


def save_comparison_results(output_folder, comparison_data, filename):
    try:
        logger.info(f"Saving comparison results to {filename}.")
        os.makedirs(output_folder, exist_ok=True)
        full_comparison_file = os.path.join(output_folder, filename)
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            for sheet_name, data in comparison_data.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        wb_full = load_workbook(full_comparison_file)
        for sheet in comparison_data.keys():
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Comparison results saved successfully to {filename}.")
    except Exception as e:
        logger.error(f"Error saving comparison results: {e}")
        raise

def main(file_previous, file_latest):
    try: 
        file_entry_previous = file_previous
        file_entry_latest = file_latest

        print(f"{file_entry_previous} + {file_entry_latest}")
        logger.info("Starting main comparison process.")
        output_folder = "ComparedResults"
        os.makedirs(output_folder, exist_ok=True)

        sheet_pr_previous, sheet_pr_latest, sheet_ADA_previous, sheet_ADA_latest, sheet_ADA_Hours_previous, sheet_ADA_Hours_latest, sheet_GOLINK_Hours_previous, sheet_GOLINK_Hours_latest, sheet_GOLINK_previous, sheet_GOLINK_latest, sheet_StandbyADAGOLINK_previous, sheet_StandbyADAGOLINK_latest, sheet_operator_previous, sheet_operator_latest, sheet_deductions_previous, sheet_deductions_latest = load_sheets(file_entry_previous, file_entry_latest)
# 1. Process the data without any filtering (full comparison)

        # 1. Compare totals
        prev_totals = calculate_totals(sheet_deductions_previous, sheet_pr_previous)
        lat_totals = calculate_totals(sheet_deductions_latest, sheet_pr_latest)
        totals_comparison_df = compare_totals(prev_totals, lat_totals)
        logger.info(f"\nTotals DF: {totals_comparison_df}")

        # Save the Hourly TTL Rev comparison results
        ttl_rev_comparison_df = compare_TTL_Rev(sheet_pr_previous, sheet_pr_latest)
        logger.info(f"Hourly TTL Rev Comparison DF: {ttl_rev_comparison_df}")

        # # Save the Per Trip Revenue comparison results
        # per_trip_revenue_comparison_df = compare_Per_Trip_Revenue(sheet_trips_previous, sheet_trips_latest)
        # logger.info(f"Per Trip Revenue Comparison DF: {per_trip_revenue_comparison_df}")

        # # 2. Compare Lift Lease
        compare_liftlease_df = compare_liftlease(sheet_pr_previous, sheet_pr_latest)
        logger.info(f"Lift Lease DF: {compare_liftlease_df}")

        # # 3. Compare Violations
        compare_violations_df = compare_violations(sheet_pr_previous, sheet_pr_latest)
        logger.info(f"Violations DF: {compare_violations_df}")

        # 4. Compare Cash Collected
        compare_cash_collected_df = compare_cash_collected(sheet_pr_previous, sheet_pr_latest)
        logger.info(f"Cash Collected DF: {compare_cash_collected_df}")

        # Save the full comparison results
        full_comparison_file = os.path.join(output_folder, "DIV3_Main_Tables.xlsx")
        excel_sheets = []
        # with pd.ExcelWriter(full_comparison_file,engine="xlsxwriter") as writer:
            # full_comparison_file = os.path.join(output_folder, "DIV3_Main_Tables.xlsx")     
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            totals_comparison_df.to_excel(writer, sheet_name="TotalInvoicePayment", index=False)
            excel_sheets.append("TotalInvoicePayment")
            ttl_rev_comparison_df.to_excel(writer, sheet_name="HourlyTTLRevComparison", index=False)
            excel_sheets.append("HourlyTTLRevComparison")
            compare_liftlease_df.to_excel(writer, sheet_name="LiftLeaseComparison", index=False)
            excel_sheets.append("LiftLeaseComparison")
            compare_violations_df.to_excel(writer, sheet_name="ViolationsComparison", index=False)
            excel_sheets.append("ViolationsComparison")
            compare_cash_collected_df.to_excel(writer, sheet_name="CashCollectedComparison", index=False)
            excel_sheets.append("CashCollectedComparison")

        # # Doperator_changes_df.to_excel(writer, sheet_name="DateComparison", index=False)

        # Apply formatting to the full comparison file
        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Main comparison process completed successfully. File saved to {full_comparison_file}.")
        
        # Save ADA comparison results
        compare_totRev_df = compare_TotRevHrs(sheet_ADA_Hours_previous, sheet_ADA_Hours_latest)
        logger.info(f"ADA Total Revenue Hours DF: {compare_totRev_df}")

        compare_PRevHrs_df = compare_PRevHrs(sheet_ADA_Hours_previous, sheet_ADA_Hours_latest)
        logger.info(f"ADA % of Revenue Hours to Forecast DF: {compare_PRevHrs_df}")

        compare_BonusHrs_df = compare_BonusHrs(sheet_ADA_Hours_previous, sheet_ADA_Hours_latest)
        logger.info(f"ADA Bonus Hours DF: {compare_BonusHrs_df}")

        compare_CoreRev_df = compare_CoreRev(sheet_ADA_Hours_previous, sheet_ADA_Hours_latest)
        logger.info(f"ADA Core Revenue DF: {compare_CoreRev_df}")

        compare_TotalEarnings_df = compare_TotalEarnings(sheet_ADA_Hours_previous, sheet_ADA_Hours_latest)
        logger.info(f"ADA Total Earnings DF: {compare_TotalEarnings_df}")

        compare_drivers_df = compare_drivers(sheet_ADA_previous, sheet_ADA_latest)
        logger.info(f"ADA Drivers DF: {compare_drivers_df}")

        # Save the full comparison results
        full_comparison_file = os.path.join(output_folder, "DIV3_ADA_Tables.xlsx")
        excel_sheets = []
        # with pd.ExcelWriter(full_comparison_file,engine="xlsxwriter") as writer:
            # full_comparison_file = os.path.join(output_folder, "DIV3_Main_Tables.xlsx")     
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            compare_totRev_df.to_excel(writer, sheet_name="TotalRevHrsComparison", index=False)
            excel_sheets.append("TotalRevHrsComparison")
            compare_PRevHrs_df.to_excel(writer, sheet_name="%RevHrsComparison", index=False)
            excel_sheets.append("%RevHrsComparison")
            compare_BonusHrs_df.to_excel(writer, sheet_name="BonusHrsComparison", index=False)
            excel_sheets.append("BonusHrsComparison")
            compare_CoreRev_df.to_excel(writer, sheet_name="CoreRevComparison", index=False)
            excel_sheets.append("CoreRevComparison")
            compare_TotalEarnings_df.to_excel(writer, sheet_name="TotEarningsComparison", index=False)
            excel_sheets.append("TotEarningsComparison")
            compare_drivers_df.to_excel(writer, sheet_name="DriversComparison", index=False)
            excel_sheets.append("DriversComparison")

        # Apply formatting to the full comparison file
        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"ADA comparison process completed successfully. File saved to {full_comparison_file}.")

        # Save GOLINK comparison results
        compare_CoreHrs_df = compare_CoreHrs(sheet_GOLINK_Hours_previous, sheet_GOLINK_Hours_latest)
        logger.info(f"GOLINK Core Hours Worked DF: {compare_CoreHrs_df}")

        compare_GPRevHrs_df = compare_PRevHrs(sheet_GOLINK_Hours_previous, sheet_GOLINK_Hours_latest)
        logger.info(f"GOLINK % Revenue Hours DF: {compare_GPRevHrs_df}")

        compare_GTotalEarnings_df = compare_TotalEarnings(sheet_GOLINK_Hours_previous, sheet_GOLINK_Hours_latest)
        logger.info(f"GOLINK Total Earnings DF: {compare_GTotalEarnings_df}")

        compare_TotDutyiolations_df = compare_TotDutyViolations(sheet_GOLINK_Hours_previous, sheet_GOLINK_Hours_latest)
        logger.info(f"GOLINK Total Duty Violation DF: {compare_TotDutyiolations_df}")

        compare_GDrivers_df = compare_drivers(sheet_GOLINK_previous, sheet_GOLINK_latest)
        logger.info(f"GOLINK Drivers DF: {compare_GDrivers_df}")

        full_comparison_file = os.path.join(output_folder, "DIV3_GOLINK_Tables.xlsx")
        excel_sheets = []
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            compare_CoreHrs_df.to_excel(writer, sheet_name="CoreHoursComparison", index=False)
            excel_sheets.append("CoreHoursComparison")
            compare_GPRevHrs_df.to_excel(writer, sheet_name="%RevHrsComparison", index=False)
            excel_sheets.append("%RevHrsComparison")
            compare_GTotalEarnings_df.to_excel(writer, sheet_name="TotEarningsComparison", index=False)
            excel_sheets.append("TotEarningsComparison")
            compare_TotDutyiolations_df.to_excel(writer, sheet_name="TotDutyViolationComparison", index=False)
            excel_sheets.append("TotDutyViolationComparison")
            compare_GDrivers_df.to_excel(writer, sheet_name="DriversComparison", index=False)
            excel_sheets.append("DriversComparison")

        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"GOLINK comparison process completed successfully. File saved to {full_comparison_file}.")

        # For ADAGOLINK comparison results
        compare_stdPayHrs_df = compare_stdPayHrs(sheet_StandbyADAGOLINK_previous, sheet_StandbyADAGOLINK_latest)
        logger.info(f"ADAGOLINK StandBy Pay Hours DF: {compare_stdPayHrs_df}")

        compare_stdExtraHrs_df = comapre_stdExtraHrs(sheet_StandbyADAGOLINK_previous, sheet_StandbyADAGOLINK_latest)
        logger.info(f"ADAGOLINK StandBy Extra Hours DF: {compare_stdExtraHrs_df}")

        compare_stdTotEarnings_df = compare_TotalEarnings(sheet_StandbyADAGOLINK_previous, sheet_StandbyADAGOLINK_latest)
        logger.info(f"ADAGOLINK Total Earnings DF: {compare_stdTotEarnings_df}")

        full_comparison_file = os.path.join(output_folder, "DIV3_ADAGOLINK_Tables.xlsx")
        excel_sheets = []
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            compare_stdPayHrs_df.to_excel(writer, sheet_name="StdByPayHrsComparison", index=False)
            excel_sheets.append("StdByPayHrsComparison")
            compare_stdExtraHrs_df.to_excel(writer, sheet_name="StByExtraHrsComparison", index=False)
            excel_sheets.append("StByExtraHrsComparison")
            compare_stdTotEarnings_df.to_excel(writer, sheet_name="TotEarningsComparison", index=False)
            excel_sheets.append("TotEarningsComparison")

        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"ADAGOLINK comparison process completed successfully. File saved to {full_comparison_file}.")

        # time.sleep(2)
        db.main(file_previous, file_latest, prev_operator, prev_date, lat_operator, lat_date)
    except Exception as e:
        logger.error(f"Error in main comparison process: {e}")
        raise



    # finally:
    #     wb_client.close()


def open_file_dialog(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsm;*.xlsx")])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(0, filename)


def create_gui():
    global file_entry_previous, file_entry_latest # Declare them as global variables

    # Set up the GUI window
    root = tk.Tk()
    root.title("Comparison Tool")

    # Create and place labels, entry boxes, and buttons
    tk.Label(root, text="Previous File:").grid(row=0, column=0, padx=10, pady=5)
    entry_previous = tk.Entry(root, width=50)
    entry_previous.grid(row=0, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_previous)).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Latest File:").grid(row=1, column=0, padx=10, pady=5)
    entry_latest = tk.Entry(root, width=50)
    entry_latest.grid(row=1, column=1, padx=10, pady=5)

    # prev = "VDP_DIV3_0310_0323 FINAL.xlsm"
    # latest = "VDP_DIV3 UPDATED TEMPLATE.xlsm"
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_latest)).grid(row=1, column=2, padx=10, pady=5)

    # Button to trigger the comparison process
    # tk.Button(root, text="Compare", command=lambda: (main(entry_previous.get(), entry_latest.get()), root.destroy())).grid(row=2, column=1, pady=20)
    tk.Button(root, text="Compare", command=lambda: handle_comparison(entry_previous.get(), entry_latest.get(), root)).grid(row=2, column=1, pady=20)
    # tk.Button(root, text="Compare", command=lambda: handle_comparison(prev, latest, root)).grid(row=2, column=1, pady=20)

    def handle_comparison(file_previous, file_latest, root):
        try:
            main(file_previous, file_latest)
        except Exception as e:
            print(f"An error occurred: {e}")
            # Check for the disconnection error and close the GUI if it happens
            if isinstance(e, OSError) and "The object invoked has disconnected" in str(e):
                print("Disconnected from Excel, closing GUI.")
                root.quit()  # This will close the Tkinter window
        finally:
            root.destroy()  # Close the window in all cases
    # Start the GUI loop
    root.mainloop()


if __name__ == "__main__":
    create_gui()

